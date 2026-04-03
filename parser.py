import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import io

def parse_excel_to_csv(file_content):
    wb = load_workbook(filename=io.BytesIO(file_content))
    ws = wb.active
    mat = [[c for c in row] for row in ws.iter_rows(values_only=True)]

    cols = list(range(3, 33))
    start1 = mat[5]
    end1   = mat[6]
    start2 = mat[7]
    end2   = mat[8]

    rows = []

    for col in cols:
        day = mat[2][col]
        if day is None:
            continue
        date = datetime(2026, 4, int(day))

        s1 = start1[col]
        e1 = end1[col]
        if isinstance(s1, (float, int)) and isinstance(e1, (float, int)):
            st = date + timedelta(hours=s1)
            et = date + timedelta(hours=e1)
            rows.append(["Dienst", st.strftime("%m/%d/%Y"), st.strftime("%H:%M"), et.strftime("%m/%d/%Y"), et.strftime("%H:%M"), "", ""])

        s2 = start2[col]
        e2 = end2[col]
        if isinstance(s2, (float, int)) and isinstance(e2, (float, int)):
            st = date + timedelta(hours=s2)
            et = date + timedelta(hours=e2)
            rows.append(["Dienst", st.strftime("%m/%d/%Y"), st.strftime("%H:%M"), et.strftime("%m/%d/%Y"), et.strftime("%H:%M"), "", ""])

    df = pd.DataFrame(rows, columns=["Subject","Start Date","Start Time","End Date","End Time","Description","Location"])
    return df.to_csv(index=False)
